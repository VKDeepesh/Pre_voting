import face_recognition
import imutils
import pickle
import time
import datetime
import cv2
import openpyxl
from imutils.video import VideoStream
from imutils.video import FPS
import os
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score
import numpy as np
import matplotlib.pyplot as plt  # Import Matplotlib for plotting graphs

# Set the path to your Excel file (attendance.xlsx) on your desktop
workbook_path = r'C:\Users\varan\OneDrive\Desktop\face attendence\attendance.xlsx'
encodingsP = r'C:\Users\varan\OneDrive\Desktop\face attendence\encodings.pickle'  # Set path to the encodings.pickle file

# Open the Excel workbook using openpyxl
workbook = openpyxl.load_workbook(workbook_path)
sheet = workbook.active

# Check if the "Number of Days" and "Time" columns exist in the header row
header_values = [cell.value for cell in sheet[1]]
if "Number of Days" not in header_values:
    sheet.cell(row=1, column=len(header_values) + 1, value="Number of Days")
if "Time" not in header_values:
    sheet.cell(row=1, column=len(header_values) + 1, value="Time")

# Initialize 'currentname' to trigger only when a new person is identified.
currentname = "unknown"

# Initialize a dictionary to store the number of days each person is recognized
num_days_recognized = {}

# Load the known faces and embeddings along with OpenCV's Haar cascade for face detection
print("[INFO] loading encodings + face detector...")
data = pickle.loads(open(encodingsP, "rb").read())  # Load the encodings from encodings.pickle

# Initialize video stream and allow the camera sensor to warm up
vs = VideoStream(src=0, framerate=10).start()
time.sleep(2.0)

# Start the FPS counter
fps = FPS().start()

# Dictionary to store the last recorded time for each face
last_recorded_time = {}

# Initialize lists to store true labels and predictions for regression accuracy calculation
true_labels = []
predictions = []
recognized_faces_over_time = []  # List to track number of faces recognized

# Loop over frames from the video file stream
while True:
    frame = vs.read()
    frame = imutils.resize(frame, width=500)
    boxes = face_recognition.face_locations(frame)
    encodings = face_recognition.face_encodings(frame, boxes)
    names = []

    for encoding in encodings:
        matches = face_recognition.compare_faces(data["encodings"], encoding)
        name = "Unknown"

        # Collect true label (1 if match, 0 if no match)
        true_label = 1 if True in matches else 0
        true_labels.append(true_label)

        if True in matches:
            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}

            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1

            name = max(counts, key=counts.get)

            if currentname != name:
                currentname = name
                print("name", currentname)

                current_time = datetime.datetime.now()

                if (
                    name not in last_recorded_time
                    or (current_time - last_recorded_time[name]).total_seconds() >= 24 * 60 * 60
                ):
                    row = [name, current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S")]
                    sheet.append(row)

                    last_recorded_time[name] = current_time

                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value == name:
                                sheet.cell(row=cell.row, column=3, value=num_days_recognized.get(name, 0) + 1)
                                sheet.cell(row=cell.row, column=4, value=current_time.strftime("%H:%M:%S"))

                if name in num_days_recognized:
                    num_days_recognized[name] += 1
                else:
                    num_days_recognized[name] = 1

        names.append(name)

    # Generate the graph and display names on the frame
    for ((top, right, bottom, left), name) in zip(boxes, names):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 225), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        cv2.putText(frame, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)

    # Track the number of recognized faces over time (frame by frame)
    recognized_faces_over_time.append(len(names))

    cv2.imshow("Facial Recognition is Running", frame)
    key = cv2.waitKey(1) & 0xFF

    if key == ord("q"):
        break

    fps.update()

# After processing, calculate accuracy using a regression model (e.g., Logistic Regression)
# Converting true_labels and predictions to arrays for regression
true_labels = np.array(true_labels)
predictions = np.array(predictions)

# You can use Logistic Regression to train and calculate the accuracy score
model = LogisticRegression()
model.fit(true_labels.reshape(-1, 1), predictions)

# Predicted values
predictions = model.predict(true_labels.reshape(-1, 1))

# Calculate accuracy (Using Accuracy Score)
accuracy = accuracy_score(true_labels, predictions)
print(f"Accuracy of the system: {accuracy * 100:.2f}%")

# Update the "Number of Days" and "Time" columns in the Excel sheet
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        name = cell.value
        if name in num_days_recognized:
            sheet.cell(row=cell.row, column=3, value=num_days_recognized[name])

# Save the changes to the workbook
workbook.save(workbook_path)

# Stop the timer and display FPS information
fps.stop()
print("[INFO] elapsed time: {:.2f}".format(fps.elapsed()))
print("[INFO] approx. FPS: {:.2f}".format(fps.fps()))

# Cleanup
cv2.destroyAllWindows()
vs.stop()

# Plot the number of recognized faces over time
plt.figure(figsize=(10, 6))
plt.plot(range(len(recognized_faces_over_time)), recognized_faces_over_time, label='Recognized Faces Over Time')
plt.xlabel('Frame Number')
plt.ylabel('Number of Recognized Faces')
plt.title('Face Recognition Trend Over Time')
plt.legend()
plt.show()
